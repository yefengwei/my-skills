#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
WKT 日报 - 填写某一天的工作内容

把「时间段 -> 叙述文本」写进对应日 sheet 的 C 列（Notes / Remarks 那一列的内容区）。
时间行由 init_week.py 生成，本脚本按「上班时间 + 序号×间隔」自行推算每行对应的时间，
不依赖 Excel 缓存值（openpyxl 读不到公式结果）。

数据格式（JSON），三选一传入：
  1) --data-file <path>
  2) --data '<json>'
  3) 管道 stdin

{
  "date": "2026-09-03",          // 可省，用 --date 指定
  "sheet": "Thursday",           // 可省，由 date 推算
  "mode": "replace",             // replace(默认) | append
  "status": "已完成",             // 可选：写入 E 列（Notes / Completion Status）
  "slots": [
     {"time": "09:00", "text": "..."},   // time = 该时段的结束时刻（模板格点）
     {"time": "10:00", "text": "..."}
  ]
}

也支持更紧凑的写法：{"slots": {"09:00": "...", "10:00": "..."}}

用法：
  python fill_day.py --date 2026-09-03 --data-file slots.json
  python fill_day.py --date 2026-09-03 --data '{"slots":[{"time":"08:30","text":"x"}]}'
  python fill_day.py --date 2026-09-03 --preview      # 只读，打印当天已有内容
  python fill_day.py --date 2026-09-03 --clear        # 清空当天内容（保留午休行）

保护规则（模板格点 9:00 起、结束点模型，行时间 = 时段结束时刻）：
  - 13:00（1:00 PM）午休行与 18:00 之后（6:00 PM 起）的行默认不会被写入，除非 --force
  - 可填工作段 8 个：9:00 / 10:00 / 11:00 / 12:00 / 14:00 / 15:00 / 16:00 / 17:00
"""

import argparse
import datetime as dt
import json
import os
import sys

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(HERE)
CONFIG_PATH = os.path.join(SKILL_DIR, "config.json")

WEEKDAY_SHEETS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def parse_hm(s):
    h, m = s.split(":")
    return int(h), int(m)


def target_path(cfg, monday):
    filename = (
        f"WKT-{cfg['staff_no']}-Daily-Work-Schedule-"
        f"{monday.year}-{cfg['staff_name']}-{monday.isoformat()}.xlsx"
    )
    return os.path.join(cfg["output_dir"], filename)


def compute_rows(cfg):
    """返回 [(时间字符串, kind, 行号)]，与 init_week.py 的模板格点保持一致。

    模板原样：G3=9:00 AM 起、60 分钟间隔，行 6 = 9:00 AM。
    行时间 = 该时段的「结束时刻」：
      - 9:00 / 10:00 / 11:00 / 12:00 行 = 上午 4 个工作段
      - 13:00（1:00 PM）行 = 午休（12:00~13:00）
      - 14:00 ~ 17:00 行 = 下午 4 个工作段
      - 18:00（6:00 PM）及以后 = 下班后的行，不填
    注意：config 的 work_start 必须与模板 G3 一致（改模板就得改 config）。
    """
    sh, sm = parse_hm(cfg["work_start"])
    eh, em = parse_hm(cfg["work_end"])
    lh, lm = parse_hm(cfg["lunch_start"])
    le_h, le_m = parse_hm(cfg["lunch_end"])
    interval = cfg["interval_min"]

    start = dt.datetime(1900, 1, 1, sh, sm)
    last = dt.datetime(1900, 1, 1, eh, em) + dt.timedelta(minutes=interval)
    lunch_s = dt.datetime(1900, 1, 1, lh, lm)
    lunch_e = dt.datetime(1900, 1, 1, le_h, le_m)
    work_end = dt.datetime(1900, 1, 1, eh, em)

    rows = []
    t = start
    idx = 0
    while t <= last:
        seg_start = t - dt.timedelta(minutes=interval)
        seg_end = t
        if seg_end > lunch_s and seg_start < lunch_e:
            kind = "lunch"
        elif seg_end <= work_end:
            kind = "work"
        else:
            kind = "off"
        rows.append((t.strftime("%H:%M"), kind, 6 + idx))
        t += dt.timedelta(minutes=interval)
        idx += 1
    return rows


def normalize_slots(slots):
    """把 {"08:30": "..."} 或 [{"time","text"}] 统一成 [(time, text)]"""
    out = []
    if isinstance(slots, dict):
        for t, text in slots.items():
            out.append((t, text))
    elif isinstance(slots, list):
        for item in slots:
            if isinstance(item, dict):
                out.append((item.get("time"), item.get("text")))
            elif isinstance(item, (list, tuple)) and len(item) == 2:
                out.append((item[0], item[1]))
    return [(t, txt) for t, txt in out if t and txt]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", help="YYYY-MM-DD，默认今天")
    ap.add_argument("--file", help="直接指定工作簿路径")
    ap.add_argument("--data-file")
    ap.add_argument("--data")
    ap.add_argument("--mode", choices=["replace", "append"], help="覆盖或追加，默认 replace")
    ap.add_argument("--status", help="写入 E 列的完成状态，如 Done / In progress")
    ap.add_argument("--preview", action="store_true", help="只读打印当天内容")
    ap.add_argument("--clear", action="store_true", help="清空当天内容")
    ap.add_argument("--force", action="store_true", help="允许覆盖午休行与下班行")
    args = ap.parse_args()

    cfg = load_config()
    d = dt.date.fromisoformat(args.date) if args.date else dt.date.today()
    monday = d - dt.timedelta(days=d.weekday())
    sheet = WEEKDAY_SHEETS[d.weekday()]

    path = args.file or target_path(cfg, monday)
    if not os.path.exists(path):
        print(json.dumps({
            "ok": False,
            "error": "week_file_not_found",
            "path": path,
            "hint": "先运行 init_week.py 创建本周工作簿",
        }, ensure_ascii=False, indent=2))
        sys.exit(2)

    rows = compute_rows(cfg)

    # ---- 预览模式 ----
    if args.preview:
        wb = load_workbook(path, data_only=False)
        ws = wb[sheet]
        preview = []
        for t, kind, r in rows:
            preview.append({
                "time": t,
                "kind": kind,
                "cell": f"C{r}",
                "text": ws[f"C{r}"].value,
                "status": ws[f"E{r}"].value,
            })
        print(json.dumps({
            "ok": True, "file": path, "sheet": sheet, "date": d.isoformat(),
            "slots": preview,
        }, ensure_ascii=False, indent=2))
        return

    wb = load_workbook(path)
    ws = wb[sheet]

    # ---- 清空模式 ----
    if args.clear:
        changed = []
        for t, kind, r in rows:
            if kind == "lunch":
                continue
            ws[f"C{r}"].value = None
            ws[f"E{r}"].value = None
            changed.append(f"C{r}")
        wb.save(path)
        print(json.dumps({"ok": True, "mode": "clear", "file": path,
                          "sheet": sheet, "cleared": changed}, ensure_ascii=False, indent=2))
        return

    # ---- 写入模式 ----
    if args.data_file:
        with open(args.data_file, "r", encoding="utf-8") as f:
            payload = json.load(f)
    elif args.data:
        payload = json.loads(args.data)
    else:
        raw = sys.stdin.read()
        if not raw.strip():
            print(json.dumps({"ok": False, "error": "no_data"}, ensure_ascii=False))
            sys.exit(1)
        payload = json.loads(raw)

    slots = normalize_slots(payload.get("slots", []))
    mode = args.mode or payload.get("mode") or "replace"
    status = args.status or payload.get("status")

    row_by_time = {t: (kind, r) for t, kind, r in rows}
    written, skipped = [], []

    for t, text in slots:
        t = t.strip()
        if t not in row_by_time:
            skipped.append({"time": t, "reason": "no_such_slot"})
            continue
        kind, r = row_by_time[t]
        if kind in ("lunch", "off") and not args.force:
            skipped.append({"time": t, "reason": f"protected_{kind}"})
            continue

        cell = ws[f"C{r}"]
        old = cell.value
        if mode == "append" and old and str(old).strip():
            if str(text) not in str(old):
                cell.value = f"{old}; {text}"
        else:
            cell.value = text
        if status:
            ws[f"E{r}"] = status
        written.append({"time": t, "cell": f"C{r}", "chars": len(str(text))})

    wb.save(path)

    print(json.dumps({
        "ok": True,
        "file": path,
        "sheet": sheet,
        "date": d.isoformat(),
        "mode": mode,
        "written": written,
        "skipped": skipped,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
