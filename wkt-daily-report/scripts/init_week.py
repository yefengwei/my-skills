#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
WKT 日报 - 周工作簿初始化（最小侵入版）

铁律（用户明确要求，2026-09-03）：
  1. 初始化只允许写一个模板位置：Data Settings!H1 = 本周一日期（日期公式链唯一源头）
  2. 模板其余一切原样保留：G3 开始时间(9:00)、I3 间隔(60 MIN)、B 列时间公式(B6~B38)、
     C5 日期公式、G 列周概览公式、各页结构 —— 绝不修改、绝不裁剪、绝不"优化"
  3. 唯一额外动作：在午休行 C 列写「午休」（内容填写，不属于结构修改）

时间格点完全来自模板：G3=9:00 AM、间隔 60 分钟 → 行6=9:00AM、行7=10:00AM ……
行时间语义为「结束点」：行时间 = 该段工作的结束时刻（用户 2026-09-03 确认）。
可填工作段由 config 的 work_start/work_end/lunch 决定（当前：9:00~17:00 共 8 段 +
13:00 午休行；6:00 PM 及以后不填）。

用法：
  python init_week.py [--date YYYY-MM-DD | --monday YYYY-MM-DD] [--dry-run]
"""

import argparse
import datetime as dt
import json
import os
import shutil
import sys

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(HERE)
CONFIG_PATH = os.path.join(SKILL_DIR, "config.json")

DAY_SHEETS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
WEEKDAY_CN = ["周一", "周二", "周三", "周四", "周五", "周六"]
SUN_SHEET = "Daily Work Schedule - SUN"
DATA_SHEET = "Data Settings"

# 模板时间行从第 6 行开始（第 5 行是表头 TIME / 日期 / Notes），G3 是第一个时间点
FIRST_ROW = 6


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def parse_hm(s):
    """'08:30' -> (8, 30)"""
    h, m = s.split(":")
    return int(h), int(m)


def build_rows(cfg):
    """模板格点行（结束点模型）：返回 [(时间串, kind, 行号)]

    - 首行 = 模板 G3（config.work_start 必须与之一致，当前 09:00），行号 6
    - kind：work = 可填工作段；lunch = 午休行；off = 下班后的行（不填）
    - 午休判定：行时段 (t-间隔, t] 与午休区间重叠
    - 只生成到 work_end + interval 为止，再往后的模板行不参与填写
    """
    sh, sm = parse_hm(cfg["work_start"])
    eh, em = parse_hm(cfg["work_end"])
    lh, lm = parse_hm(cfg["lunch_start"])
    le_h, le_m = parse_hm(cfg["lunch_end"])
    interval = cfg["interval_min"]

    start = dt.datetime(1900, 1, 1, sh, sm)
    last = dt.datetime(1900, 1, 1, eh, em) + dt.timedelta(minutes=interval)
    lunch_s = dt.datetime(1900, 1, 1, lh, lm)
    lunch_e = dt.datetime(1900, 1, 1, le_h, le_m)
    work_end = dt.datetime(1900, 1, 1, eh, em)

    rows = []
    t = start
    idx = 0
    while t <= last:
        seg_start = t - dt.timedelta(minutes=interval)
        seg_end = t
        if seg_end > lunch_s and seg_start < lunch_e:
            kind = "lunch"
        elif seg_end <= work_end:
            kind = "work"
        else:
            kind = "off"
        rows.append((t.strftime("%H:%M"), kind, FIRST_ROW + idx))
        t += dt.timedelta(minutes=interval)
        idx += 1
    return rows


def hm(t):
    return t.strftime("%H:%M")


def week_type(monday, big_monday):
    """根据大周基准周一推算：返回 'big' / 'small' / None"""
    if not big_monday:
        return None
    base = dt.date.fromisoformat(big_monday)
    delta_weeks = (monday - base).days // 7
    if delta_weeks < 0:
        return None
    return "big" if delta_weeks % 2 == 0 else "small"


def resolve_template(cfg):
    """模板路径兜底：config 指向的文件不存在时，在输出目录里找最新的 *模版*.xlsx。

    公司会不定期换模板文件名（如把「工号」占位符换成真实编号），
    有这个兜底就不用每次都改 config。
    """
    tpl = cfg["template"]
    if os.path.exists(tpl):
        return tpl
    outdir = cfg["output_dir"]
    cands = [
        os.path.join(outdir, f)
        for f in os.listdir(outdir)
        if f.endswith(".xlsx") and "模版" in f
    ]
    if not cands:
        raise FileNotFoundError(f"模板不存在: {tpl}，且 {outdir} 下未找到 *模版*.xlsx")
    found = max(cands, key=os.path.getmtime)
    print(json.dumps({"note": "config.template 不存在，已自动切换到最新模板", "template": found},
                     ensure_ascii=False), file=sys.stderr)
    return found


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", help="任意日期 YYYY-MM-DD，默认今天")
    ap.add_argument("--monday", help="直接指定周一日期 YYYY-MM-DD")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    cfg = load_config()
    rows = build_rows(cfg)

    if args.monday:
        monday = dt.date.fromisoformat(args.monday)
    else:
        d = dt.date.fromisoformat(args.date) if args.date else dt.date.today()
        monday = d - dt.timedelta(days=d.weekday())

    filename = (
        f"WKT-{cfg['staff_no']}-Daily-Work-Schedule-"
        f"{monday.year}-{cfg['staff_name']}-{monday.isoformat()}.xlsx"
    )
    outdir = cfg["output_dir"]
    os.makedirs(outdir, exist_ok=True)
    target = os.path.join(outdir, filename)
    template = resolve_template(cfg)

    created = False
    if not os.path.exists(target):
        if args.dry_run:
            print(json.dumps({"dry_run": True, "would_create": target}, ensure_ascii=False, indent=2))
            return
        shutil.copy2(template, target)
        created = True

    if args.dry_run:
        print(json.dumps({"dry_run": True, "exists": target, "created": created}, ensure_ascii=False, indent=2))
        return

    wb = load_workbook(target)

    # ===== 唯一允许写入的模板位置：周一日期 =====
    ds = wb[DATA_SHEET]
    ds["H1"] = dt.datetime(monday.year, monday.month, monday.day)
    ds["H1"].number_format = "yyyy-mm-dd"

    # ===== 唯一的内容动作：午休行写「午休」（仅 6 个日页；SUN 页属于"其余位置"，不动） =====
    lunch_rows = [r for t, k, r in rows if k == "lunch"]
    for name in DAY_SHEETS:
        ws = wb[name]
        for r in lunch_rows:
            if str(ws[f"C{r}"].value or "").strip() == "":
                ws[f"C{r}"] = "午休"

    # ===== 输出映射（供 fill_day 与 AI 参考，不写文件） =====
    result_days = {}
    for i, name in enumerate(DAY_SHEETS):
        day_date = monday + dt.timedelta(days=i)
        result_days[name] = {
            "sheet": name,
            "date": day_date.isoformat(),
            "weekday_cn": WEEKDAY_CN[i],
            "rows": {t: f"C{r}" for t, k, r in rows if k in ("work", "lunch")},
            "work_slots": [t for t, k, r in rows if k == "work"],
            "protected": [t for t, k, r in rows if k != "work"],
        }

    wb.save(target)

    out = {
        "file": target,
        "created": created,
        "monday": monday.isoformat(),
        "sunday": (monday + dt.timedelta(days=6)).isoformat(),
        "week_type": week_type(monday, cfg.get("big_week_monday")),
        "grid": "模板原样：G3 开始时间 + 60 分钟间隔，行时间=时段结束点",
        "work_start": cfg["work_start"],
        "work_end": cfg["work_end"],
        "lunch": f"{cfg['lunch_start']}-{cfg['lunch_end']}",
        "days": result_days,
    }
    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
